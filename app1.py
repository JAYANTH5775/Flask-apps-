from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)


@app.route('/')
def index():
    return render_template("index.html")

@app.route('/submit', methods=['POST'])
def feedback_form():
    if request.method == 'POST':

        visited = request.form.get('visited', 'na')
        reason = request.form.get('reason', 'na')
        needed = request.form.get('needed', 'na')
        looking_for = request.form.get('looking_for', 'na')
        easy = request.form.get('easy', 'na')
        impression = request.form.get('impression', 'nskdjv')
        with open('data.txt', 'w') as file:
            file.write(f"{visited}, {reason}, {needed}, {looking_for}, {easy}, {impression}\n")

        try:
            wb = load_workbook('template.xlsx')
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['Visited', 'Primary reason', 'Found what needed', 'Looking for', 'Easy to find info',
                       'Overall impression'])

        # Check if the header row exists and add it if necessary
        if not ws['A1'].value:
            header_row = ['Visited', 'Primary reason', 'Found what needed', 'Looking for', 'Easy to find info',
                          'Overall impression']
            for col_num, value in enumerate(header_row, start=1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = value

        ws.append([visited, reason, needed, looking_for, easy, impression])
        wb.save('template.xlsx')

        return render_template('success.html')

    return render_template('index.html')


# @app.route('/')
# def index():
#     return render_template("index.html")
#
# @app.route('/submit', methods=['POST'])
# def feedback_form():
#     if request.method == 'POST':
#
#         visited = request.form.get('visited','na')
#         reason = request.form.get('reason','na')
#         needed = request.form.get('needed','na')
#         looking_for = request.form.get('looking_for','na')
#         easy = request.form.get('easy','na')
#         impression = request.form.get('impression','nskdjv')
#
#         try:
#             wb = load_workbook('template.xlsx')
#             ws = wb.active
#         except FileNotFoundError:
#             wb = Workbook()
#             ws = wb.active
#             ws.append(['Visited', 'Primary reason', 'Found what needed', 'Looking for', 'Easy to find info', 'Overall impression'])
#         ws.append([visited, reason, needed, looking_for, easy, impression])
#
#         wb.save('template.xlsx')
#
#         return render_template('success.html')
#
#     return render_template('index.html')

if __name__ == "__main__":
    app.run(debug=True)
