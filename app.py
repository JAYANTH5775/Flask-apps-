from flask import Flask  , render_template , request


from openpyxl import load_workbook


app = Flask(__name__)

@app.route('/')

def index():
    return render_template("index.html")
@app.route('/', methods=['POST'])
def feedback_form():
    if request.method == 'POST':
        # Retrieve the form data
        visited = request.form['visited']
        reason = request.form['reason']
        needed = request.form['needed']
        looking_for = request.form['looking_for']
        easy = request.form['easy']
        impression = request.form['impression']
        # Add the form data to the Excel sheet
        # ...
        # Return a message indicating successful submission
        return "Feedback submitted successfully"
    # Render the feedback form template if the request method is GET
    return render_template('success.html')

    # Import the openpyxl module
    from openpyxl import Workbook
    # Create a new workbook
    wb = Workbook()
    # Select the active worksheet
    ws = wb.active
    # Set the column headings
    ws.append(['Visited', 'Primary reason', 'Found what needed', 'Looking for', 'Easy to find info', 'Overall impression'])
    # Open the workbook
    wb = load_workbook('template.xlsx')
    # Select the active worksheet
    ws = wb.active
    # Append the form data to the sheet
    ws.append([visited, reason, needed, looking_for, easy, impression])
    # Save the workbook
    wb.save('template.xlsx')

# def process_data():
#     reason = request.form['reason']
#
#
#
#
#
#     wb  = load_workbook('template.xlsx')
#
#     ws = wb.active
#     ws.column_dimensions['A'].width = 30
#
#     rows = ws.max_row +1
#     ws.cell(row = 1 , column=1, value =reason)
#     # ws.cell(row = row, column=2 , value=data)
#
#     wb.save(f'{reason}.xlsx')
#
#     return f"reason:{reason}"

if __name__ == "__main__":
    app.run(debug=True)